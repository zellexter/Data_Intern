from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    LineChart,
    Reference,
    PieChart,
    ProjectedPieChart,
)
from openpyxl.chart.label import DataLabelList
import os
from openpyxl.chart.axis import DateAxis
import pandas as pd
import numpy as np
from collections import Counter
from datetime import date

# COMMIT COMMENTS
# generalized draw_piechart() to use min/max col/row to find data
# draw_piechart() now passes in chart name as output ws name
# removed usecol= argument from import_data(), redundent as columns are found by str not index. all columns from input file are imported to df.
# generalized for loops in main() so col_name points toward needed columns in df
# format_data_linechart() passes in is_month with a default value of False.
# generalize variables in format_data_piechart() 

# draw_linechart() passes in x_axis and y_axis in function and chart title 
#   when defining the class method need to pass in self as first argument. not needed when you call the method.


# DONE visualize data into date line chart https://openpyxl.readthedocs.io/en/stable/charts/line.html
# DONE line 62 linechart_fw_creation_date.y_axis.crossAx ????
# DONE split error in line 26
# DONE make efficient with ws
# DONE rename file
# TODO try out df.aggregate
# DONE def visulaize(col_name, is_month, chart_type)
# DONE edit functions to include 'if is_month'
# DONE save into different worksheets
#   DONE main() add new arg into for loop that defines ws variable as prepare_ws()
#   DONE function prepare_ws() that creates new ws (like prepare_wb) with specified title (related to col_name)
#   DONE revisit prepare_wb() and edit out obsolete code
#   DONE format_data() wb to ws, appends all formatted data into specified ws instead of wb
#   DONE draw_line_chart() wb to ws, take data from specified ws instead of wb
#   DONE save_chart() wb to ws, save chart into specified ws instead of wb
# DONE generalize into class
# DONE generalize format_data() to exclude pulling specific columns
#   DONE pass in which columns and where to put data
#   DONE pass in type of data and how to treat
# TODO create net type data visualization piechart
#   TODO new function to add unoriginal data to df
#       ??? how to approach: make new function or pass in original_data=True to format_data function
#   DONE create general pie chart for mode types
# DONE generalize reference columns, size of current data (len)
#   DONE generalize piechart.title
# DONE generalize for loops in main function so that you can pass in the needed columns of data from df
# TODO create VisualizationType() Class with all draw_chart functions and format data functions*** WAIT FOR DAD 
# TODO/IN PROGRESS LINE 212 format_data()
#   generalize format data function to pass in type of chart, combine pie and line format data functions
# TODO generalize draw chart function to pass in type of chart
# TODO generalize main() for loops to pass in type of chart


# STEP BY STEP
# DONE get raw data 
# DONE aggregate data into usable format
#   DONE line chart by day)
#       format into two cols, date and how many fw created on each date
#   DONE line chart by month)
#       format date data into mm/yyyy - testing on testpage.py
#       format data into list within list containing [date, count]
#       append into ws
#       chart
# visualization:
#   DONE line chart for firewall creation date by day
#   bar chart ^^^ by month
#   multilevel donut chart for mode and nettype within mode

### note: put visualizations and data onto same sheet

# basic_user_registration_data  self disclosed data
# user_viewing_behavior_data  uid, time, like, share, comment, topic
# user_follow_behavior_data   uid, following, follow_by
# heuristic_user_data         

class DataVisualization():
    '''TODO'''

    def __init__(self, input_file, output_file) -> None:
        self.input_file = input_file
        self.output_file = output_file
        self.wb = None

    def main(self):
        '''Main Function
            Assigns dataframe into usable variable
            Prepares a workbook
            Loops through specified data columns and creates unique worksheet where visualization is stored
            Saves workbook'''
        
        df = self.import_data()
        
        self.wb = self.prepare_wb()

        # create linechart
        for col_name, output_worksheet, is_month, x_axis, y_axis in [
            ('FwCreateDate', 'Daily Chart', False, 'Date', 'Firewalls Created'),
            ('FwCreateMonth', 'Monthly Chart', True, 'Date', 'Firewalls Created'),
            ]:
            ws = self.prepare_ws(output_worksheet)
            self.format_data_linechart(df, ws, col_name, is_month)
            self.draw_linechart(ws, x_axis, y_axis, output_worksheet)


        # create piechart
        for col_name, output_worksheet in [
            ('Mode', 'Mode Piechart'),
            ('CloudSec_NetType', 'CloudSec NetType Piechart'),
            ('NetSec_NetType', 'Netsec NetType Piechart'),
            # IN ORDER TO DO THESE^^ need to add data to df
            # df.NetTypeOfCloudSec = func(df.NetType)
        ]:
            ws = self.prepare_ws(output_worksheet)
            self.format_data_piechart(df, ws, col_name)
            self.draw_piechart(ws, chart_title=output_worksheet)

        self.save_wb()
        

    def import_data(self):
        '''This function reads an excel sheet and imports specific columns into a dataframe.
        Creates by month data column from date data. Returns the dataframe.'''

        # import xlsx file, Data sheet, columns for mode, fwcreatedate, and nettype
        dict_df = pd.read_excel(io=self.input_file, sheet_name=['Data']) 
        df = dict_df['Data']

        # generalize TODO
        df['FwCreateMonth'] = df['FwCreateDate'].str.slice(0,7)
        df['CloudSec_NetType'] = df.apply(lambda row: row['NetType'] if row['Mode'] == 'CloudSec' else np.nan, axis=1)
        df['NetSec_NetType'] = df.apply(lambda row: row['NetType'] if row['Mode'] == 'NetSec' else np.nan, axis=1)

        return df


    def prepare_wb(self):
        '''This function creates a new workbook'''
        wb = Workbook()
        wb.remove(wb['Sheet'])
        return wb


    def prepare_ws(self, ws_name): # pass in col_name as name for ws
        '''This function creates a new worksheet named output_worksheet in workbook'''
        ws = self.wb.create_sheet(ws_name)
        return ws


    def format_data_linechart(self, df, ws, col_name, is_month=False):
        '''This function formats date data for a linechart'''

        data = df.loc[:, col_name]

        # creates sorted_count_date list within list [date, count], adds header in formatted_date_data
        count_data = Counter(list(data))
        count_data.pop(np.nan, None) # pop np.nan from data, if nonexistent then do nothing

        # definition of when is_month is True
        is_month = list(count_data.keys())[0].count('-') == 1
        # what to do when is_month is True
        if is_month:
            count_data = {
                (k + '-01') : v # what we are doing
                for k, v in count_data.items() # for which items
                if k is not np.nan # conditional # use is not for np.nan because we don't know if unknowns == unknowns, but we DO know if something IS NOT unknown
                }
        # list_count_data = count_data.items()
            
        sorted_count_data = sorted([[date(*[int(item) for item in k.split('-')]),v] #what we are doing to the items
                                for k,v in count_data.items() #where we find the items from
                                if k is not np.nan] #conditional filter
                                )
        count_data_headers = [[col_name,'Count']]
        formatted_date_data = count_data_headers + sorted_count_data
        # return formatted_date_data

        # appends formatted_date_data to ws2 only if ws2 is empty (to stop it from appending again every time the code is run)
        if ws['A1'].value == None: # check ws2 is empty
            for row in formatted_date_data:
                ws.append(row)

    def format_data_piechart(self, df, ws, col_name):
        '''This function formats data for a piechart'''

        # create column for data
        data = df.loc[:, col_name]

        # creates count_date
        count_data = Counter(list(data))
        count_data.pop(np.nan, None) # pop np.nan from data, if nonexistent then do nothing

        # total variable to assist in converting to percentage
        # Counter({'NetSec': 943, 'CloudSec': 723})
        total_count = count_data.total()

        # list comp to sort data types into usable format
        sorted_count_data = sorted([[k,v/total_count]
                                    for k,v in count_data.items()]
                                   )
        count_data_headers = [[col_name, 'Count']]
        formatted_data = count_data_headers + sorted_count_data
        # return formatted_mode_data

        # appends formatted_mode_data to ws2 only if ws2 is empty (to stop it from appending again every time the code is run)
        if ws['A1'].value == None: # check ws2 is empty
            for row_num, row_data in enumerate(formatted_data, start=2):
                ws.append(row_data)
                # convert to percentage, specific to piechart
                ws.cell(row=row_num, column=2).number_format = '0.0%'


    # def _format_general(self):

    # ??? Would this be better as Class format_data(), with general code and then specific functions for each type of chart
    def format_data(self, df, ws, col_name, charttype):
        '''Generalized function for formatting data.
            Pass in charttype to format in specific way'''
        
        ### GENERAL
        # self._format_general()  # private method, vs public

        # create column for data
        data = df.loc[:, col_name]

        # creates count_date
        count_data = Counter(list(data))
        count_data.pop(np.nan, None) # pop np.nan from data, if nonexistent then do nothing
    
        ### CHART SPECIFIC

        # if charttype == linechart:

        # if charttype == piechart:

    def draw_linechart(self, ws, x_axis, y_axis, chart_title):
        '''This function draws a line chart and saves it into a cell'''

        # Line Chart
        linechart = LineChart()
        linechart.title = chart_title
        linechart.style = 12
        linechart.y_axis.title = y_axis
        linechart.y_axis.crossAx = 500 # wth does this even mean
        linechart.x_axis = DateAxis(crossAx=100)
        linechart.x_axis.number_format = 'd-mmm'
        linechart.x_axis.majorTimeUnit = 'days'
        linechart.x_axis.title = x_axis
        
        # where to reference data from
        data = Reference(ws,
                        min_col=2,
                        min_row=2,
                        max_col=2,
                        max_row=ws.max_row,
                        )
        dates = Reference(ws,
                        min_col=1,
                        min_row=2,
                        max_col=1,
                        max_row=ws.max_row,
                        )
        linechart.add_data(data, titles_from_data=True)
        linechart.set_categories(dates)

        ws.add_chart(linechart, 'C1')

    def draw_piechart(self, ws, chart_title):
        '''This function draws a pie chart and saves it into a cell'''
        
        # creating pie chart
        piechart = PieChart()
        labels = Reference(ws, min_col=ws.min_column, min_row=3, max_row=ws.max_row)
        data = Reference(ws, min_col=ws.max_column, min_row=2, max_row=ws.max_row)
        piechart.add_data(data, titles_from_data=True)
        piechart.set_categories(labels)
        piechart.title = chart_title

        # showing data labels as percentage
        piechart.dataLabels = DataLabelList()
        piechart.dataLabels.showVal = True

        # add chart to ws
        ws.add_chart(piechart, 'C1')

    def save_wb(self):
        '''This function saves the excel file'''
        self.wb.save(self.output_file)


# if DataVisualization() is run on the original file, use specified filepath and save_file.
if __name__ == '__main__':
    filepath = r'C:\Users\miche\OneDrive\Documents\CodingwDad\Data_Intern\py-xl-chart.xlsx'
    save_file = os.path.join('Data_Intern', 'data_intern.xlsx')
    DataVisualization(input_file=filepath, output_file=save_file).main()
        # class() create an instance of the class
        # example : Eevee = Canine(gen=female, col=brown, size=m)
        #           Cobra = Canine(gen=male, col=tribrown, size=xs)











